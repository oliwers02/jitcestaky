"""Exporty: XLSX (openpyxl), PDF (fpdf2), CSV (MRP/Money S3/Pohoda), XML."""
from __future__ import annotations

import datetime as dt
import io
import re
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core import calc

# ---------------------------------------------------------------- pomocné ---

CESTY_STLPCE = [
    ("datum", "Dátum"),
    ("zamestnanec", "Zamestnanec"),
    ("trasa", "Trasa"),
    ("ucel_cesty", "Účel"),
    ("vozidlo", "Vozidlo"),
    ("vzdialenost_km", "Km"),
    ("vypocitana_phm_nahrada", "PHM náhrada (€)"),
    ("vypocitana_zakladna_nahrada", "Základná náhrada (€)"),
    ("vypocitane_stravne", "Stravné"),
    ("stravne_mena", "Mena stravného"),
    ("vedlajsie_vydavky_eur", "Vedľajšie výdavky (€)"),
    ("nahrada_spolu", "Spolu (€)"),
]


def trips_to_df(trips: list[dict], covered_days: set | None = None) -> pd.DataFrame:
    """Jednodňové cesty -> DataFrame so slovenskými stĺpcami pre export.

    Ak je dátum jednodňovej cesty pokrytý viacdňovou cestou (covered_days),
    jej stravné sa do hárku Cesty NEdoplní (=0) a o túto sumu sa zníži „Spolu",
    aby sa stravné nezdvojilo s hárkom Diéty (žije len v diétach).
    """
    covered_days = covered_days or set()
    rows = []
    for t in trips:
        v_multidni = (t.get("employee_id"), str(t.get("datum"))) in covered_days
        stravne = 0.0 if v_multidni else float(t.get("vypocitane_stravne") or 0)
        stravne_orig_eur = calc._trip_stravne_eur(t) if v_multidni else 0.0
        spolu = round(float(t.get("nahrada_spolu") or 0) - stravne_orig_eur, 2)
        rows.append({
            "Dátum": t.get("datum", ""),
            "Zamestnanec": t.get("zamestnanec", ""),
            "Trasa": f"{t.get('miesto_zaciatku') or ''} – {t.get('ciel_cesty') or ''}",
            "Účel": t.get("ucel_cesty", ""),
            "Vozidlo": t.get("vozidlo", ""),
            "Doprava": calc.TYP_DOPRAVY.get(t.get("typ_dopravy", ""), t.get("typ_dopravy", "")),
            "Krajina": t.get("cielova_krajina", "SK"),
            "Km": float(t.get("vzdialenost_km") or 0),
            "PHM náhrada (EUR)": float(t.get("vypocitana_phm_nahrada") or 0),
            "Základná náhrada (EUR)": float(t.get("vypocitana_zakladna_nahrada") or 0),
            "Stravné": stravne,
            "Mena stravného": t.get("stravne_mena", "EUR"),
            "Stravné v rámci viacdňovej": "áno (v Diétach)" if v_multidni else "—",
            "Vreckové": float(t.get("vypocitane_vreckove") or 0),
            "Vedľajšie výdavky (EUR)": float(t.get("vedlajsie_vydavky_eur") or 0),
            "Spolu (EUR)": spolu,
        })
    return pd.DataFrame(rows)


def diety_to_df(business_trips: list[dict]) -> pd.DataFrame:
    """Denný rozpis diét všetkých (zahraničných aj tuzemských) viacdňových ciest."""
    from core import db
    rows = []
    for bt in business_trips:
        dni = db.fetch_all("business_trip_days", "business_trip_id = ?",
                           (bt["id"],), order="datum")
        krajina = bt.get("cielova_krajina") or "SK"
        if not bt.get("zahranicna"):
            krajina = "SK"
        for d in dni:
            _, mena, pasmo = calc.stravne_pre_hodiny(
                float(d.get("pocet_hodin") or 0), krajina,
                calc.get_rates(d["datum"]))
            jedla = [n for n, k in [("raňajky −25 %", "ranajky_poskytnute"),
                                    ("obed −40 %", "obed_poskytnuty"),
                                    ("večera −35 %", "vecera_poskytnuta")]
                     if d.get(k)]
            rows.append({
                "Cesta": bt.get("nazov", ""),
                "Deň": d["datum"],
                "Typ dňa": calc.TYP_DNA.get(d["typ_dna"], d["typ_dna"]),
                "Hodiny": float(d.get("pocet_hodin") or 0),
                "Krajina": krajina,
                "Mena": mena,
                "Pásmo": pasmo,
                "Krátenie jedál": ", ".join(jedla) if jedla else "—",
                "Stravné": float(d.get("stravne_den_eur") or 0),
            })
    return pd.DataFrame(rows)


# ------------------------------------------------------------------- XLSX ---

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_THIN = Border(*[Side(style="thin", color="BBBBBB")] * 4)


def _write_sheet(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for c in ws[1]:
        c.fill, c.font = _HDR_FILL, _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    for _, row in df.iterrows():
        ws.append(list(row))
    for j, col in enumerate(df.columns, start=1):
        width = max([len(str(col))] + [len(str(v)) for v in df[col].head(50)]) + 3
        ws.column_dimensions[get_column_letter(j)].width = min(width, 45)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(df.columns)):
        for c in row:
            c.border = _THIN
            if isinstance(c.value, float):
                c.number_format = "#,##0.00"


def export_xlsx(trips_df: pd.DataFrame, diety_df: pd.DataFrame,
                suhrn: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Súhrn"
    ws.append(["Cestovné príkazy — súhrn"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for k, v in suhrn.items():
        ws.append([k, v])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 24

    ws2 = wb.create_sheet("Cesty")
    if not trips_df.empty:
        _write_sheet(ws2, trips_df)
    else:
        ws2.append(["Žiadne jednodňové cesty vo zvolenom období."])

    ws3 = wb.create_sheet("Diéty")
    if not diety_df.empty:
        _write_sheet(ws3, diety_df)
    else:
        ws3.append(["Žiadne viacdňové cesty vo zvolenom období."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -------------------------------------------------------------------- CSV ---

def export_csv(df: pd.DataFrame, sep: str = ";",
               encoding: str = "cp1250") -> bytes:
    """CSV na import do účtovného softvéru (MRP, Money S3, Pohoda).

    Pre SK účtovné programy je zvyčajne vhodný oddeľovač ';' a kódovanie
    Windows-1250; alternatívne UTF-8 (s BOM kvôli Excelu).
    """
    if encoding.lower() in ("utf-8", "utf8"):
        text = df.to_csv(index=False, sep=sep)
        return ("﻿" + text).encode("utf-8")
    return df.to_csv(index=False, sep=sep).encode(encoding, errors="replace")


# -------------------------------------------------------------------- XML ---

def export_xml(trips: list[dict], business_trips: list[dict],
               dni_map: dict[int, list[dict]] | None = None) -> bytes:
    root = ET.Element("cestovne_prikazy",
                      {"vygenerovane": dt.datetime.now().isoformat(timespec="seconds"),
                       "legislativa": "zákon č. 283/2002 Z.z.", "rok_sadzieb": "2026"})
    for t in trips:
        c = ET.SubElement(root, "cesta", {"id": str(t.get("id", "")), "typ": "jednodnova"})
        for tag, key in [("datum", "datum"), ("zamestnanec", "zamestnanec"),
                         ("vozidlo", "vozidlo"), ("miesto_zaciatku", "miesto_zaciatku"),
                         ("ciel", "ciel_cesty"), ("ucel", "ucel_cesty"),
                         ("typ_dopravy", "typ_dopravy"), ("krajina", "cielova_krajina")]:
            ET.SubElement(c, tag).text = str(t.get(key) or "")
        s = ET.SubElement(c, "sumy", {"mena": "EUR"})
        ET.SubElement(s, "vzdialenost_km").text = f"{float(t.get('vzdialenost_km') or 0):.1f}"
        ET.SubElement(s, "phm_nahrada").text = f"{float(t.get('vypocitana_phm_nahrada') or 0):.2f}"
        ET.SubElement(s, "zakladna_nahrada").text = f"{float(t.get('vypocitana_zakladna_nahrada') or 0):.2f}"
        ET.SubElement(s, "stravne", {"mena": t.get("stravne_mena", "EUR")}).text = \
            f"{float(t.get('vypocitane_stravne') or 0):.2f}"
        ET.SubElement(s, "vedlajsie_vydavky").text = f"{float(t.get('vedlajsie_vydavky_eur') or 0):.2f}"
        ET.SubElement(s, "spolu").text = f"{float(t.get('nahrada_spolu') or 0):.2f}"

    for bt in business_trips:
        c = ET.SubElement(root, "cesta", {"id": str(bt.get("id", "")), "typ": "viacdnova"})
        for tag, key in [("nazov", "nazov"), ("zamestnanec", "zamestnanec"),
                         ("ucel", "ucel"), ("datum_zaciatku", "datum_zaciatku"),
                         ("datum_konca", "datum_konca"), ("krajina", "cielova_krajina"),
                         ("status", "status")]:
            ET.SubElement(c, tag).text = str(bt.get(key) or "")
        s = ET.SubElement(c, "sumy", {"mena": "EUR"})
        ET.SubElement(s, "ubytovanie").text = f"{float(bt.get('ubytovanie_eur') or 0):.2f}"
        ET.SubElement(s, "kilometrovne").text = f"{float(bt.get('kilometrovne_eur') or 0):.2f}"
        ET.SubElement(s, "navstevy_rodiny").text = f"{float(bt.get('navstevy_rodiny_eur') or 0):.2f}"
        if dni_map and bt["id"] in dni_map:
            diety = ET.SubElement(c, "diety")
            for d in dni_map[bt["id"]]:
                de = ET.SubElement(diety, "den", {"datum": d["datum"],
                                                  "typ": d["typ_dna"]})
                ET.SubElement(de, "hodiny").text = f"{float(d.get('pocet_hodin') or 0):.2f}"
                ET.SubElement(de, "ranajky").text = str(int(d.get("ranajky_poskytnute") or 0))
                ET.SubElement(de, "obed").text = str(int(d.get("obed_poskytnuty") or 0))
                ET.SubElement(de, "vecera").text = str(int(d.get("vecera_poskytnuta") or 0))
                ET.SubElement(de, "stravne").text = f"{float(d.get('stravne_den_eur') or 0):.2f}"
    ET.indent(root)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


# -------------------------------------------------------------------- PDF ---

# Font priložený v repozitári má prednosť — Unicode (diakritika) funguje
# na každej platforme vrátane Streamlit Cloud, kde systémové fonty nemusia byť.
_BUNDLED_FONTS = Path(__file__).resolve().parent.parent / "assets" / "fonts"
_FONT_CANDIDATES = [
    (str(_BUNDLED_FONTS / "DejaVuSans.ttf"),
     str(_BUNDLED_FONTS / "DejaVuSans-Bold.ttf")),
    ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
     "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
    ("/System/Library/Fonts/Supplemental/Arial.ttf",
     "/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
]

# Podpisový obrázok priložený v repozitári (ak existuje, vloží sa do PDF).
_SIGNATURE = Path(__file__).resolve().parent.parent / "assets" / "signatures" / "podpis.png"

# Náhrady bežných „nelatin-1" znakov pre prípad fallbacku na core font.
_ASCII_MAP = {
    "—": "-", "–": "-", "‑": "-", "−": "-",
    "“": '"', "”": '"', "„": '"', "‟": '"', "’": "'", "‘": "'",
    "…": "...", "€": "EUR", "→": "->", "×": "x", "•": "-", "\xa0": " ",
}


def _ascii_fallback(text: str) -> str:
    """Bezpečný prepis do latin-1 pre prípad, že nie je Unicode font.
    Odstráni diakritiku, prepíše pomlčky/úvodzovky/€ a zvyšok nahradí '?'."""
    s = str(text)
    for src, dst in _ASCII_MAP.items():
        s = s.replace(src, dst)
    s = "".join(c for c in unicodedata.normalize("NFKD", s)
                if not unicodedata.combining(c))
    return s.encode("latin-1", "replace").decode("latin-1")


class _PDF(FPDF):
    """PDF s unicode fontom, ak je dostupný; inak fallback bez diakritiky."""

    def __init__(self):
        super().__init__(format="A4")
        self.set_auto_page_break(True, margin=18)
        self.unicode_ok = False
        for reg, bold in _FONT_CANDIDATES:
            if Path(reg).exists() and Path(bold).exists():
                self.add_font("App", "", reg)
                self.add_font("App", "B", bold)
                self.unicode_ok = True
                break
        self.family = "App" if self.unicode_ok else "helvetica"

    def t(self, text: str) -> str:
        return str(text) if self.unicode_ok else _ascii_fallback(text)

    def h1(self, text: str):
        self.set_font(self.family, "B", 15)
        self.cell(0, 9, self.t(text), new_x="LMARGIN", new_y="NEXT", align="C")
        self.ln(2)

    def h2(self, text: str):
        self.set_font(self.family, "B", 11)
        self.cell(0, 7, self.t(text), new_x="LMARGIN", new_y="NEXT")

    def kv(self, k: str, v: str):
        col = 58  # šírka stĺpca pre popis (mm)
        label = self.t(k)
        self.set_font(self.family, "B", 9.5)
        if self.get_string_width(label) > col - 2:
            # dlhý popis -> na samostatnom riadku, hodnota odsadená pod ním
            self.cell(0, 6, label, new_x="LMARGIN", new_y="NEXT")
            self.set_font(self.family, "", 9.5)
            self.set_x(self.l_margin + col)
            self.multi_cell(0, 6, self.t(v), new_x="LMARGIN", new_y="NEXT")
        else:
            self.cell(col, 6, label, border=0)
            self.set_font(self.family, "", 9.5)
            self.multi_cell(0, 6, self.t(v), new_x="LMARGIN", new_y="NEXT")

    def table(self, headers: list[str], rows: list[list], widths: list[float]):
        self.set_font(self.family, "B", 8.5)
        self.set_fill_color(225, 232, 240)
        for h, w in zip(headers, widths):
            self.cell(w, 6.5, self.t(h), border=1, fill=True, align="C")
        self.ln()
        self.set_font(self.family, "", 8.5)
        for row in rows:
            for v, w in zip(row, widths):
                align = "R" if isinstance(v, (int, float)) else "L"
                txt = f"{v:.2f}" if isinstance(v, float) else str(v)
                self.cell(w, 6, self.t(txt), border=1, align=align)
            self.ln()


def _hlavicka(pdf: _PDF, profil: dict):
    pdf.set_font(pdf.family, "B", 10)
    pdf.cell(0, 5.5, pdf.t(profil.get("nazov") or ""), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(pdf.family, "", 9)
    riadky = [profil.get("adresa") or ""]
    ids = " | ".join(f"{k}: {profil[p]}" for k, p in
                     [("IČO", "ico"), ("DIČ", "dic"), ("IČ DPH", "icdph")]
                     if profil.get(p))
    if ids:
        riadky.append(ids)
    for r in riadky:
        if r:
            pdf.cell(0, 5, pdf.t(r), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.line(pdf.l_margin, pdf.get_y(), 210 - pdf.r_margin, pdf.get_y())
    pdf.ln(4)


def _format_datum(datum) -> str:
    """ISO 'YYYY-MM-DD' -> 'DD.MM.YYYY' (deň podpisu na dokumente)."""
    try:
        return dt.date.fromisoformat(str(datum)).strftime("%d.%m.%Y")
    except (ValueError, TypeError):
        return str(datum or "")


def _podpisy(pdf: _PDF, datum=None):
    """Podpisové bloky pre obe strany. Ak existuje podpisový obrázok, vloží sa
    nad obe čiary; dátum (deň, kedy sa má dokument podpísať podľa zákona) sa
    doplní pod oba popisy."""
    pdf.ln(18)
    y = pdf.get_y()
    if y > 245:
        pdf.add_page()
        y = pdf.get_y() + 12
    # podpisový obrázok tesne nad čiarou (vľavo aj vpravo)
    if _SIGNATURE.exists():
        w = 46.0
        h = w * 74 / 293  # pomer orezaného podpisu
        pdf.image(str(_SIGNATURE), x=55 - w / 2, y=y - h - 0.5, w=w)
        pdf.image(str(_SIGNATURE), x=155 - w / 2, y=y - h - 0.5, w=w)
    pdf.set_draw_color(0, 0, 0)
    pdf.line(25, y, 85, y)
    pdf.line(125, y, 185, y)
    pdf.set_y(y + 1)
    pdf.set_font(pdf.family, "", 9)
    pdf.cell(90, 5, pdf.t("podpis zamestnanca"), align="C")
    pdf.cell(0, 5, pdf.t("podpis schvaľujúceho"),
             new_x="LMARGIN", new_y="NEXT", align="C")
    if datum:
        pdf.set_font(pdf.family, "", 8.5)
        d = _format_datum(datum)
        pdf.cell(90, 5, pdf.t(f"Dátum: {d}"), align="C")
        pdf.cell(0, 5, pdf.t(f"Dátum: {d}"),
                 new_x="LMARGIN", new_y="NEXT", align="C")


def pdf_cestovny_prikaz(bt: dict, zamestnanec: dict, profil: dict) -> bytes:
    """PDF „Príkaz na pracovnú cestu" (pred cestou)."""
    pdf = _PDF()
    pdf.add_page()
    _hlavicka(pdf, profil)
    pdf.h1("CESTOVNÝ PRÍKAZ")
    pdf.set_font(pdf.family, "", 9)
    pdf.cell(0, 5, pdf.t("podľa zákona č. 283/2002 Z. z. o cestovných náhradách"),
             new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)
    pdf.kv("Zamestnanec:", zamestnanec.get("meno_priezvisko", ""))
    pdf.kv("Pozícia:", zamestnanec.get("pozicia") or "—")
    pdf.kv("Bydlisko (nástup):", zamestnanec.get("adresa_bydliska") or "—")
    pdf.ln(2)
    pdf.kv("Názov cesty:", bt.get("nazov") or "—")
    pdf.kv("Účel cesty:", bt.get("ucel") or "—")
    pdf.kv("Miesto/krajina:", calc.KRAJINY.get(bt.get("cielova_krajina", "SK"),
                                               bt.get("cielova_krajina", "SK")))
    pdf.kv("Začiatok cesty:", f"{bt.get('datum_zaciatku', '')} o {bt.get('cas_odchodu_prvy_den') or '—'}")
    pdf.kv("Koniec cesty:", f"{bt.get('datum_konca', '')} o {bt.get('cas_navratu_posledny_den') or '—'}")
    pdf.kv("Spôsob dopravy:", "súkromné cestné motorové vozidlo")
    pdf.kv("Poznámky:", bt.get("poznamky") or "—")
    pdf.ln(3)
    suhlas = "ÁNO" if bt.get("suhlas_zamestnanca") else "NIE"
    pdf.kv("Súhlas zamestnanca s vyslaním (§3):", suhlas)
    pdf.ln(2)
    pdf.set_font(pdf.family, "", 8)
    pdf.multi_cell(0, 4.5, pdf.t(
        "Zamestnancovi patrí náhrada preukázaných cestovných výdavkov, výdavkov za "
        "ubytovanie, stravné a náhrada preukázaných potrebných vedľajších výdavkov "
        "podľa zákona č. 283/2002 Z. z."))
    # cestovný príkaz sa vystavuje a podpisuje pred nástupom na cestu
    _podpisy(pdf, datum=bt.get("datum_zaciatku"))
    return bytes(pdf.output())


def pdf_vyuctovanie(bt: dict, zamestnanec: dict, profil: dict,
                    dni: list[dict], suhrn: dict) -> bytes:
    """PDF „Vyúčtovanie pracovnej cesty" s denným rozpisom stravného."""
    pdf = _PDF()
    pdf.add_page()
    _hlavicka(pdf, profil)
    pdf.h1("VYÚČTOVANIE PRACOVNEJ CESTY")
    pdf.ln(1)
    pdf.kv("Zamestnanec:", zamestnanec.get("meno_priezvisko", ""))
    pdf.kv("Cesta:", f"{bt.get('nazov') or ''} ({bt.get('ucel') or ''})")
    pdf.kv("Termín:", f"{bt.get('datum_zaciatku')} {bt.get('cas_odchodu_prvy_den') or ''}"
                      f" – {bt.get('datum_konca')} {bt.get('cas_navratu_posledny_den') or ''}")
    pdf.kv("Krajina:", calc.KRAJINY.get(bt.get("cielova_krajina", "SK"),
                                        bt.get("cielova_krajina", "SK")))
    pdf.kv("Status:", calc.STATUSY.get(bt.get("status", ""), bt.get("status", "")))
    pdf.ln(3)

    pdf.h2("Denný rozpis stravného")
    mena = suhrn.get("mena", "EUR")
    rows = []
    for d in dni:
        jedla = "".join([("R" if d.get("ranajky_poskytnute") else "–"),
                         ("O" if d.get("obed_poskytnuty") else "–"),
                         ("V" if d.get("vecera_poskytnuta") else "–")])
        rows.append([d["datum"], calc.TYP_DNA.get(d["typ_dna"], d["typ_dna"]),
                     float(d.get("pocet_hodin") or 0), jedla,
                     float(d.get("stravne_den_eur") or 0)])
    pdf.table(["Dátum", "Typ dňa", "Hodiny", "Jedlá (R/O/V)", f"Stravné ({mena})"],
              rows, [32, 42, 24, 38, 38])
    pdf.set_font(pdf.family, "", 7.5)
    pdf.cell(0, 5, pdf.t("Krátenie za poskytnuté jedlá: raňajky −25 %, obed −40 %, večera −35 %."),
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.h2("Rekapitulácia náhrad")
    rk = [
        ["Stravné spolu", f"{suhrn['stravne']:.2f} {mena}"],
        ["Vreckové", f"{suhrn['vreckove']:.2f} {mena}"],
        ["Ubytovanie", f"{float(bt.get('ubytovanie_eur') or 0):.2f} EUR"],
        ["Kilometrovné (PHM + základná náhrada)", f"{float(bt.get('kilometrovne_eur') or 0):.2f} EUR"],
        ["Návštevy rodiny", f"{float(bt.get('navstevy_rodiny_eur') or 0):.2f} EUR"],
    ]
    if suhrn.get("spolu_eur") is not None:
        rk.append(["SPOLU (v EUR)", f"{suhrn['spolu_eur']:.2f} EUR"])
    else:
        rk.append(["SPOLU", f"{suhrn['ostatne_eur']:.2f} EUR + {suhrn['stravne']:.2f} {mena}"
                            " (kurz CZK/EUR nie je zadaný v Nastaveniach)"])
    pdf.table(["Položka", "Suma"], rk, [110, 64])
    # vyúčtovanie sa podpisuje po skončení cesty (deň návratu)
    _podpisy(pdf, datum=bt.get("datum_konca"))
    return bytes(pdf.output())


def pdf_prehlad(trips_df: pd.DataFrame, diety_df: pd.DataFrame,
                suhrn: dict, profil: dict, nazov: str) -> bytes:
    """Súhrnný PDF report pre vybrané obdobie (stránka Export)."""
    pdf = _PDF()
    pdf.add_page()
    _hlavicka(pdf, profil)
    pdf.h1(nazov)
    pdf.ln(1)
    for k, v in suhrn.items():
        pdf.kv(f"{k}:", str(v))
    pdf.ln(3)
    if not trips_df.empty:
        pdf.h2("Jednodňové cesty")
        rows = [[r["Dátum"], (r["Trasa"][:34] + "…") if len(str(r["Trasa"])) > 35 else r["Trasa"],
                 float(r["Km"]), float(r["PHM náhrada (EUR)"]),
                 float(r["Základná náhrada (EUR)"]), float(r["Stravné"]),
                 float(r["Spolu (EUR)"])]
                for _, r in trips_df.iterrows()]
        pdf.table(["Dátum", "Trasa", "Km", "PHM €", "Zákl. €", "Stravné", "Spolu €"],
                  rows, [22, 62, 16, 19, 19, 19, 19])
        pdf.ln(4)
    if not diety_df.empty:
        pdf.h2("Diéty (viacdňové cesty)")
        rows = [[r["Deň"], (str(r["Cesta"])[:28] + "…") if len(str(r["Cesta"])) > 29 else r["Cesta"],
                 r["Krajina"], r["Pásmo"], float(r["Hodiny"]),
                 f"{r['Stravné']:.2f} {r['Mena']}"]
                for _, r in diety_df.iterrows()]
        pdf.table(["Deň", "Cesta", "Krajina", "Pásmo", "Hodiny", "Stravné"],
                  rows, [22, 50, 20, 30, 18, 36])
    return bytes(pdf.output())


# ------------------------------------------- PDF jednodňových ciest ---------

def pdf_prikaz_jednodnova(trip: dict, zamestnanec: dict, profil: dict,
                          vozidlo_str: str = "") -> bytes:
    """PDF „Cestovný príkaz" pre jednodňovú cestu."""
    pdf = _PDF()
    pdf.add_page()
    _hlavicka(pdf, profil)
    pdf.h1("CESTOVNÝ PRÍKAZ")
    pdf.set_font(pdf.family, "", 9)
    pdf.cell(0, 5, pdf.t("jednodňová cesta — podľa zákona č. 283/2002 Z. z."),
             new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)
    pdf.kv("Zamestnanec:", zamestnanec.get("meno_priezvisko", ""))
    pdf.kv("Pozícia:", zamestnanec.get("pozicia") or "—")
    pdf.kv("Bydlisko (nástup):", zamestnanec.get("adresa_bydliska") or "—")
    pdf.ln(2)
    pdf.kv("Dátum cesty:", _format_datum(trip.get("datum")))
    pdf.kv("Čas:", f"{trip.get('cas_odchodu') or '—'} – {trip.get('cas_prichodu') or '—'}")
    pdf.kv("Trasa:", f"{trip.get('miesto_zaciatku') or ''} – {trip.get('ciel_cesty') or ''}")
    pdf.kv("Účel cesty:", trip.get("ucel_cesty") or "—")
    pdf.kv("Miesto/krajina:", calc.KRAJINY.get(trip.get("cielova_krajina", "SK"),
                                               trip.get("cielova_krajina", "SK")))
    doprava = "súkromné cestné motorové vozidlo" \
        if trip.get("typ_dopravy") == "sukromne_auto" else "verejná doprava"
    pdf.kv("Spôsob dopravy:", doprava + (f" ({vozidlo_str})" if vozidlo_str else ""))
    pdf.ln(3)
    suhlas = "ÁNO" if trip.get("suhlas_zamestnanca") else "NIE"
    pdf.kv("Súhlas zamestnanca s vyslaním (§3):", suhlas)
    _podpisy(pdf, datum=trip.get("datum"))
    return bytes(pdf.output())


def pdf_vyuctovanie_jednodnova(trip: dict, zamestnanec: dict, profil: dict,
                               vozidlo_str: str = "") -> bytes:
    """PDF „Vyúčtovanie pracovnej cesty" pre jednodňovú cestu."""
    pdf = _PDF()
    pdf.add_page()
    _hlavicka(pdf, profil)
    pdf.h1("VYÚČTOVANIE PRACOVNEJ CESTY")
    pdf.set_font(pdf.family, "", 9)
    pdf.cell(0, 5, pdf.t("jednodňová cesta"), new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(3)
    pdf.kv("Zamestnanec:", zamestnanec.get("meno_priezvisko", ""))
    pdf.kv("Dátum cesty:", _format_datum(trip.get("datum")))
    pdf.kv("Čas:", f"{trip.get('cas_odchodu') or '—'} – {trip.get('cas_prichodu') or '—'}")
    pdf.kv("Trasa:", f"{trip.get('miesto_zaciatku') or ''} – {trip.get('ciel_cesty') or ''}")
    pdf.kv("Účel cesty:", trip.get("ucel_cesty") or "—")
    if vozidlo_str:
        pdf.kv("Vozidlo:", vozidlo_str)
    pdf.ln(2)

    hod = calc.hodiny_trvania(trip.get("cas_odchodu"), trip.get("cas_prichodu"))
    krajina = trip.get("cielova_krajina", "SK") if trip.get("zahranicna") else "SK"
    _, mena_str, pasmo = calc.stravne_pre_hodiny(
        hod, krajina, calc.get_rates(trip.get("datum")))
    mena = trip.get("stravne_mena") or mena_str
    km = float(trip.get("vzdialenost_km") or 0)

    pdf.h2("Rekapitulácia náhrad")
    rk = [
        ["Vzdialenosť (tam a späť)", f"{km:.1f} km"],
        ["PHM náhrada", f"{float(trip.get('vypocitana_phm_nahrada') or 0):.2f} EUR"],
        ["Základná náhrada za km", f"{float(trip.get('vypocitana_zakladna_nahrada') or 0):.2f} EUR"],
        [f"Stravné ({pasmo}, {hod:.1f} h)",
         f"{float(trip.get('vypocitane_stravne') or 0):.2f} {mena}"],
        ["Vedľajšie výdavky", f"{float(trip.get('vedlajsie_vydavky_eur') or 0):.2f} EUR"],
    ]
    if float(trip.get("vypocitane_vreckove") or 0) > 0:
        rk.append(["Vreckové", f"{float(trip.get('vypocitane_vreckove') or 0):.2f} {mena}"])
    rk.append(["SPOLU", f"{float(trip.get('nahrada_spolu') or 0):.2f} EUR"])
    pdf.table(["Položka", "Suma"], rk, [110, 64])
    _podpisy(pdf, datum=trip.get("datum"))
    return bytes(pdf.output())


# --------------------------------------------------- ZIP balíček PDF --------

def _slug(text: str, maxlen: int = 40) -> str:
    s = _ascii_fallback(text).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return (s[:maxlen] or "cesta")


def build_zip(items: list[tuple[str, bytes]]) -> bytes:
    """Zabalí (nazov_suboru, obsah) položky do ZIP archívu."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in items:
            z.writestr(name, data)
    return buf.getvalue()
